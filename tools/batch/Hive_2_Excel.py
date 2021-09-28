import difflib

import openpyxl
import os
import re

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from HqlParse import HqlParse

# 清洗字段得到字段名
def clean_column(columns):
    new_columns = []
    for column in columns:
        column = re.sub(r',\s*?\w+\(.*\)\s*?OVER\([A-Za-z0-9_ \n,]*?\)\s+(?:(as)?\s*?)(\w+)', r',\2', column,
                          flags=re.M | re.I)
        column = re.sub(r'\s*?case.*?as\s+(\w+)', r' \1', column, flags=re.M | re.I | re.S)
        column = re.sub(r'(\w+\s+?)as(\s+?\w+)', r'\2', column, flags=re.M | re.S | re.I)
        column = re.sub(r'^\w+\.(\w+).*',r'\1',column,flags=re.M | re.S | re.I)
        column = re.sub(r',?(\w+)\s*.*',r'\1',column,flags=re.M | re.S | re.I)
        new_columns.append(column)
    return new_columns

# 拿到字段的表别名
def get_columnwithalias(columns):
    new_column_dic = {}
    for column in columns:
        alias = re.sub(r'\s*,?(\w+)\..*',r'\1',column,flags=re.M | re.S | re.I)
        column = re.sub(r',\s*?\w+\(.*\)\s*?OVER\([A-Za-z0-9_ \n,]*?\)\s+(?:(as)?\s*?)(\w+)', r',\2', column,
                        flags=re.M | re.I)
        column = re.sub(r'\s*?case.*?as\s+(\w+)', r' \1', column, flags=re.M | re.I | re.S)
        column = re.sub(r'(\w+\s+?)as(\s+?\w+)', r'\2', column, flags=re.M | re.S | re.I)
        column = re.sub(r'\s*,?(\w+)\.\s*?(\w+).*', r'\2', column, flags=re.M | re.S | re.I)
        new_column_dic[column] = alias
    return new_column_dic

#判断相似度的方法，用到了difflib库
def get_equal_rate_1(str1, str2):
   return difflib.SequenceMatcher(None, str1, str2).quick_ratio()


def hive_2_excel(sql):
    filedir = 'D:/target/excel'
    filename = ''
    # 创建新表格
    #new_workbook = openpyxl.Workbook()
    # 读取表格数据
    new_workbook = load_workbook('D:\Excel_python\数据API详细设计文档.xlsx')
    new_sheet = new_workbook['Sheet3']
    parse = HqlParse(sql)
    tables = parse.tables[0]
    table_name = tables['table_name'][tables['table_name'].rfind('.')+1:]
    new_sheet.title = table_name
    new_sheet['A2'].value = table_name
    all_table_info = {}
    # 主键
    target_pk = re.findall(r'--\s*?@\s*?Primary Key[：:]\s*?([\w,]+)\s*?', sql, re.S | re.M | re.I)
    if target_pk:
        target_pk = target_pk[0].split(',')
    # 表的中文名称
    comment = HqlParse.extract_table_comment(sql)
    if comment:
        new_sheet['B2'].value = comment

    if isinstance(target_pk,list):
        index = 2
        for pk in target_pk:
            new_sheet['C'+ str(index)].value = pk
            index += 1
    elif target_pk:
        new_sheet['C2'].value = target_pk[0]
    # 新表格的数据从第二行开始写入
    new_row_idx = 2
    # 模型表字段字典
    column_dic = {}
    # for column in definitions:
    #     columns.append(' {name!s:12} {definition}'.format(
    #         name=column[0], definition=' '.join(str(t) for t in column[1:])))
    thin = Side(border_style="thin", color="000000")
    for column in tables['definitions']:
        # 字段名
        new_sheet['D' + str(new_row_idx)].value = column[0].value
        # 字段类型
        new_sheet['E' + str(new_row_idx)].value = column[1].value
        # 字段注释
        new_sheet['F' + str(new_row_idx)].value = column[3].value
        new_sheet['L' + str(new_row_idx)].value = column[3].value
        column_dic[column[0].value] = {'row_idx':new_row_idx,'type':column[1].value,'notes':column[3].value}
        new_row_idx += 1

    insert_info = parse.insert_info[0]
    # 目标表数据
    target_tname = insert_info['target_table_name']
    target_t_columns = insert_info['target_columns']
    target_columns_dic = get_columnwithalias(target_t_columns)


    # ----- 处理主表数据
    if insert_info['sub_tables']:
        sub_table_list = []
        main_table = insert_info['sub_tables'][0]['table']
        main_tname = main_table['table_name'][main_table['table_name'].rfind('.')+1:]

        main_columns = clean_column(main_table['columns'])

        main_alias_name = main_table['table_alias']
        all_table_info[main_alias_name] = main_tname
        # 写主表数据
        # for main_column in main_columns:
        #     if main_column in target_columns_dic.keys():
        #         if main_column in column_dic.keys() and target_columns_dic[main_column] == main_alias_name:
        #             row_idx = column_dic[main_column]['row_idx']
        #             # 来源表表名
        #             new_sheet['H' + str(row_idx)].value = main_tname
        #             # 来源表字段名称
        #             new_sheet['J' + str(row_idx)].value = main_column
        #             # 来源表字段类型
        #             new_sheet['K' + str(row_idx)].value = column_dic[main_column]['type']
        for key in target_columns_dic.keys():
            if  target_columns_dic[key] == main_alias_name and key in column_dic.keys():
                row_idx = column_dic[key]['row_idx']
                # 来源表表名
                new_sheet['H' + str(row_idx)].value = main_tname
                # 来源表字段名称
                new_sheet['J' + str(row_idx)].value = key
                # 来源表字段类型
                new_sheet['K' + str(row_idx)].value = column_dic[key]['type']



        # 处理子表
        # ----- 获取子表数据
        def get_table_struct(table_info):
            table_res = table_info['res']
            table_info = table_info['table']
            table_columns = table_info['columns']
            table_name = None
            if not (table_columns) or len(table_columns) == 0:
                return None
            if not table_info['table_name']:
                try:
                    table_name = table_info['subtable'][0]['table_name']
                except:
                    pass
            else:
                table_name = table_info['table_name'][table_info['table_name'].rfind('.')+1:]

            table_alias = table_info['table_alias']
            all_table_info[table_alias] = table_name
            table_edge_list = []
            table_res_etl = []
            for res in table_res:
                table_edge = re.findall(r'(\w+?)\.[\w =]+?(\w+?)\.', res, re.S | re.I | re.M)  # 表的连线
                res = re.sub(r'(.*?)(\s+?--.*)', r'\1', res, re.S | re.I | re.M)
                table_res_etl.append(res)
                if table_edge:
                    table_edge_list.append(table_edge[0])
            table_columns = [re.sub(r'[\s|\n|\'|`]', '', colum) for colum in table_columns]  # 清洗
            table_columns = clean_column(table_columns)
            return [table_columns, table_name, table_alias, table_res_etl, table_edge_list]
        # 处理所有的子表
        sub_list = insert_info['sub_tables'][1:]

        for index in range(len(sub_list)):
            # print(get_table_struct(sub_list[index]))
            sub_table_list.append(get_table_struct(sub_list[index]))
        try:
            sub_table_list.remove(None)
        except:
            pass

        for sub_table in sub_table_list:
            sub_columns = sub_table[0]
            sub_name    = sub_table[1]
            sub_alias   = sub_table[2]
            sub_res     = sub_table[3]
            table_edge_list = sub_table[4]
            # 处理 sub_res
            sub_res_str = ''
            for res in sub_res:
                table_edge = re.findall(r'(\w+?)\.[\w =]+?(\w+?)\.', res, re.S | re.I | re.M)  # 表的连线
                table_edge = table_edge[0] if table_edge  else None
                if not table_edge:
                    continue
                table_columns = re.findall(r'\s*\w+?\.(\w+)[ =]+\w+\.(\w+)', res, re.S | re.I | re.M)[0]  # 表的字段
                if len(table_edge) == 2:
                    if table_edge[0] in all_table_info.keys() and table_edge[1] in all_table_info.keys():
                        sub_res_str = """{0}.{1} = \n{2}.{3}\n""".format(all_table_info[table_edge[0]]
                                                             ,table_columns[0]
                                                             ,all_table_info[table_edge[1]]
                                                             ,table_columns[1])

            flag = 0
            for key in target_columns_dic.keys():
                if target_columns_dic[key] == sub_alias and key in column_dic.keys():
                    row_idx = column_dic[key]['row_idx']
                    # 来源表表名
                    new_sheet['H' + str(row_idx)].value = sub_name
                    # 来源表字段名称
                    new_sheet['J' + str(row_idx)].value = key
                    # 来源表字段类型
                    new_sheet['K' + str(row_idx)].value = column_dic[key]['type']
                    # 关联键
                    if not flag:
                        new_sheet['G' + str(row_idx)].value = sub_res_str
                        flag = 1


            # for sub_column in sub_columns:
            #     if sub_column in target_columns_dic.keys():
            #         if sub_column in column_dic.keys() and target_columns_dic[sub_column] == sub_alias:
            #             row_idx = column_dic[sub_column]['row_idx']
            #             # 来源表表名
            #             new_sheet['H' + str(row_idx)].value = sub_name
            #             # 来源表字段名称
            #             new_sheet['J' + str(row_idx)].value = sub_column
            #             # 来源表字段类型
            #             new_sheet['K' + str(row_idx)].value = column_dic[sub_column]['type']
            #             # 关联键
            #             if not flag:
            #                 new_sheet['G' + str(row_idx)].value = sub_res_str
            #                 flag = 1



    filename = '数据API详细设计文档-{0}V1.0.xlsx'.format(table_name)
    # 保存修改
    new_workbook.template = False
    new_workbook.save(filedir + '/' + filename)
    return filedir,filename

